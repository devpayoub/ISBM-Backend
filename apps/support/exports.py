"""PDF (reportlab) and Excel (openpyxl) export for SAV tickets.

Both libraries were already in requirements/base.txt but unused anywhere
else in the codebase — this is the first real usage of either.
"""
import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

COLUMNS = [
    "N° ticket", "Machine", "Criticité", "Statut", "Créé le", "Description",
    "Fournisseur", "Temps d'arrêt (min)", "Durée intervention (min)",
    "Pièces remplacées", "Coût intervention",
]


def _rows(queryset):
    for t in queryset:
        closure = getattr(t, "closure", None)
        yield [
            t.ticket_number,
            t.machine.code,
            t.get_criticality_display(),
            t.get_status_display(),
            t.created_at.strftime("%Y-%m-%d %H:%M"),
            (t.description or "")[:120],
            t.assigned_supplier.full_name if t.assigned_supplier_id else "",
            closure.total_downtime_min if closure else "",
            closure.intervention_duration_min if closure and closure.intervention_duration_min is not None else "",
            (closure.parts_replaced or "")[:120] if closure else "",
            closure.intervention_cost if closure and closure.intervention_cost is not None else "",
        ]


def export_tickets_pdf(queryset) -> HttpResponse:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph("Tickets SAV", styles["Title"]), Spacer(1, 12)]

    data = [COLUMNS] + list(_rows(queryset))
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
    ]))
    elements.append(table)
    doc.build(elements)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="tickets_sav.pdf"'
    return response


def export_tickets_excel(queryset) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets SAV"
    ws.append(COLUMNS)
    for row in _rows(queryset):
        ws.append(row)
    for i, _ in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="tickets_sav.xlsx"'
    return response
